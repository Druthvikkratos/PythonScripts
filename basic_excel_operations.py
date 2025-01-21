#load exisitng file
import openpyxl # type: ignore

"""
wb = openpyxl.load_workbook('ganesh.xlsx')
ws = wb.active
print("Excel file loaded successfully.")
"""

#create a excel file
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "SampleSheet"

ws.append(["ID", "Name", "Age", "Score"])
ws.append([1, "Alice", 21, 85])
ws.append([2, "Bob", 23, 90])

#update cell value
ws["B2"]="Alice Johonson"

#update score for all rows
"""
for row in ws.iter_rows(min_row=2, max_col=4, values_only=False):
    row[3].value += 5
"""

#insert new row
"""
ws.insert_rows(4)
ws["A4"] = 3
ws["B4"] = "Charlie"
ws["C4"] = 22
ws["D4"] = 90
"""
"""
ws.delete_cols(3)
"""

#lists sheenames
#sheetname = wb.sheetnames

#create sheets
wb.create_sheet(index=1, title="new sheet")
print("sheet is created")

sheet_to_remove = wb["new sheet"]
wb.remove(sheet_to_remove)

wb.save("example.xlsx")
print("Excel file created successfully.")


sheetname = wb.sheetnames
print(sheetname)